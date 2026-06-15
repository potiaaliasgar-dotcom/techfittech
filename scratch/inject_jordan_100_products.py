import json
import re
import openpyxl

APP_JS = "/Users/batman/Desktop/techfittech/public/assets/app.js"
EXCEL_PATH = "/Users/batman/Desktop/techfittech/BH Price List 2026 (BH, Jordan, Bendis pilates).xlsx"

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
sheet = wb['Jordan']

print("Loaded Excel sheet. Max rows:", sheet.max_row)

categories_map = [
    # (keyword_in_name, exclude_keyword, category_name, series_name, section_name, image_slug)
    # Racks (Match racks first to avoid matching the items they hold!)
    ("dumbbell rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "dumbbell-rack"),
    ("barbell rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "barbell-rack"),
    ("bar rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "barbell-rack"),
    ("bar holder", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "barbell-rack"),
    ("toaster rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "plate-rack"),
    ("plate rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "plate-rack"),
    ("kettlebell rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "kettlebell-rack"),
    ("storage rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "dumbbell-rack"),
    ("power rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "dumbbell-rack"),
    ("squat rack", None, "Storage Racks", "ACCESSORIES & MATS", "Free Weights", "dumbbell-rack"),
    
    # Dumbbells
    ("custom urethane dumbbells", "set", "Urethane Dumbbells", "DUMBBELLS", "Free Weights", "urethane-dumbbells"),
    ("custom urethane dumbbell set", None, "Urethane Dumbbell Sets", "DUMBBELLS", "Free Weights", "urethane-dumbbells"),
    ("hex rubber dumbbells", "set", "Hex Rubber Dumbbells", "DUMBBELLS", "Free Weights", "hex-rubber-dumbbells"),
    ("hex rubber dumbbell set", None, "Hex Dumbbell Sets", "DUMBBELLS", "Free Weights", "hex-rubber-dumbbells"),
    ("custom hybrid dumbbells", None, "Chrome Dumbbells", "DUMBBELLS", "Free Weights", "chrome-dumbbells"),
    ("chrome dumbbell set", None, "Chrome Dumbbell Sets", "DUMBBELLS", "Free Weights", "chrome-dumbbells"),
    ("neoprene studio dumbbells", "set", "Studio Dumbbells", "DUMBBELLS", "Free Weights", "neoprene-dumbbells"),
    
    # Kettlebells
    ("neoprene kettlebell", "set", "Neoprene Kettlebells", "KETTLEBELLS", "Free Weights", "neoprene-kettlebells"),
    ("neoprene kettlebell set", None, "Kettlebell Sets", "KETTLEBELLS", "Free Weights", "neoprene-kettlebells"),
    ("competition kettlebell", "set", "Competition Kettlebells", "KETTLEBELLS", "Free Weights", "neoprene-kettlebells"),
    ("cast iron kettlebell", "set", "Cast Iron Kettlebells", "KETTLEBELLS", "Free Weights", "neoprene-kettlebells"),
    
    # Barbells & Plates
    ("urethane dual grip olympic plate", None, "Urethane Plates", "BARBELLS & PLATES", "Free Weights", "urethane-plates"),
    ("urethane tri-grip olympic plate", None, "Urethane Plates", "BARBELLS & PLATES", "Free Weights", "urethane-plates"),
    ("rubber bumper plate", None, "Bumper Plates", "BARBELLS & PLATES", "Free Weights", "rubber-bumper-plates"),
    ("urethane bumper plate", None, "Bumper Plates", "BARBELLS & PLATES", "Free Weights", "rubber-bumper-plates"),
    ("aluminium training bar", None, "Olympic Bars", "BARBELLS & PLATES", "Free Weights", "aluminium-bar"),
    ("olympic bar", None, "Olympic Bars", "BARBELLS & PLATES", "Free Weights", "aluminium-bar"),
    ("urethane barbell set", None, "Urethane Barbells", "BARBELLS & PLATES", "Free Weights", "aluminium-bar"),
    ("rubber barbell set", None, "Rubber Barbells", "BARBELLS & PLATES", "Free Weights", "aluminium-bar"),
    ("rubber barbell", "set", "Rubber Barbells", "BARBELLS & PLATES", "Free Weights", "aluminium-bar"),
    ("cambered bar", None, "Olympic Bars", "BARBELLS & PLATES", "Free Weights", "aluminium-bar"),
    
    # Functional
    ("double grip medicine ball", None, "Medicine Balls", "FUNCTIONAL", "Functional Training", "double-grip-medicine-balls"),
    ("medicine ball", "double grip", "Medicine Balls", "FUNCTIONAL", "Functional Training", "double-grip-medicine-balls"),
    ("flexi bag", None, "Power Bags", "FUNCTIONAL", "Functional Training", "sandbags"),
    ("sandbag extreme", None, "Power Bags", "FUNCTIONAL", "Functional Training", "sandbags"),
    ("slam ball", None, "Slam Balls", "FUNCTIONAL", "Functional Training", "slam-balls"),
    ("rope trainer", None, "Rope Trainers", "FUNCTIONAL", "Functional Training", "rope-trainer"),
    ("battle rope", None, "Battle Ropes", "FUNCTIONAL", "Functional Training", "rope-trainer"),
    ("blazepod", None, "BlazePod Kits", "FUNCTIONAL", "Functional Training", "blazepod"),
    ("punch bag", None, "Boxing Gear", "FUNCTIONAL", "Functional Training", "punch-bag"),
    ("water punch bag", None, "Boxing Gear", "FUNCTIONAL", "Functional Training", "punch-bag"),
    
    # Accessories & Mats
    ("airex corona", None, "AIREX Mats", "ACCESSORIES & MATS", "Free Weights", "airex-corona-mat"),
    ("airex coronella", None, "AIREX Mats", "ACCESSORIES & MATS", "Free Weights", "airex-corona-mat"),
    ("airex fitline", None, "AIREX Mats", "ACCESSORIES & MATS", "Free Weights", "airex-corona-mat"),
    ("airex balance", None, "AIREX Balance Pads", "ACCESSORIES & MATS", "Free Weights", "airex-corona-mat"),
    ("studio barbell set", None, "Studio Pump Sets", "ACCESSORIES & MATS", "Free Weights", "ignite-pump-set"),
    ("pump x urethane studio barbell", None, "Studio Pump Sets", "ACCESSORIES & MATS", "Free Weights", "ignite-pump-set"),
]

extracted_products = []
seen_slugs = set()

for i in range(4, sheet.max_row + 1):
    sku = sheet.cell(row=i, column=1).value
    product = sheet.cell(row=i, column=3).value
    price = sheet.cell(row=i, column=4).value
    
    if not sku or not product or price is None or str(price).startswith('#'):
        continue
        
    product_clean = product.strip()
    lines = product_clean.split('\n')
    name = lines[0].strip()
    name = ' '.join(name.split())
    
    desc = ""
    if len(lines) > 1:
        desc = " ".join([l.strip() for l in lines[1:] if l.strip()])
    else:
        desc = f"Premium {name} by Jordan Fitness, designed for commercial health clubs, boutique gyms, and professional training facilities."
    
    name_lower = name.lower()
    
    matched_cat = None
    for kw, ex_kw, cat_name, sr_name, sec_name, img_slug in categories_map:
        if kw in name_lower:
            if ex_kw and ex_kw in name_lower:
                continue
            matched_cat = (cat_name, sr_name, sec_name, img_slug)
            break
            
    if not matched_cat:
        # Clean word-based matching for fallback cases
        if re.search(r'\bmat\b|\bmats\b', name_lower):
            matched_cat = ("AIREX Mats", "ACCESSORIES & MATS", "Free Weights", "airex-corona-mat")
        elif "kettlebell" in name_lower:
            if "rack" in name_lower or "shelf" in name_lower:
                matched_cat = ("Storage Racks", "ACCESSORIES & MATS", "Free Weights", "kettlebell-rack")
            else:
                matched_cat = ("Kettlebells", "KETTLEBELLS", "Free Weights", "neoprene-kettlebells")
        elif "dumbbell" in name_lower:
            if "rack" in name_lower:
                matched_cat = ("Storage Racks", "ACCESSORIES & MATS", "Free Weights", "dumbbell-rack")
            else:
                matched_cat = ("Dumbbells", "DUMBBELLS", "Free Weights", "urethane-dumbbells")
        elif "plate" in name_lower:
            if "rack" in name_lower or "toaster" in name_lower:
                matched_cat = ("Storage Racks", "ACCESSORIES & MATS", "Free Weights", "plate-rack")
            else:
                matched_cat = ("Urethane Plates", "BARBELLS & PLATES", "Free Weights", "urethane-plates")
        elif "rack" in name_lower or "holder" in name_lower or "shelf" in name_lower:
            if "bar" in name_lower:
                matched_cat = ("Storage Racks", "ACCESSORIES & MATS", "Free Weights", "barbell-rack")
            else:
                matched_cat = ("Storage Racks", "ACCESSORIES & MATS", "Free Weights", "dumbbell-rack")
        elif "punch bag" in name_lower or "boxing" in name_lower or "bag" in name_lower:
            matched_cat = ("Boxing Gear", "FUNCTIONAL", "Functional Training", "punch-bag")
        elif "blazepod" in name_lower:
            matched_cat = ("BlazePod Kits", "FUNCTIONAL", "Functional Training", "blazepod")
        else:
            continue
            
    cat_name, sr_name, sec_name, img_slug = matched_cat
    
    base_slug = re.sub(r'[^a-z0-9\-]', '', name.lower().replace(' ', '-').replace('/', '-'))
    base_slug = re.sub(r'\-+', '-', base_slug).strip('-')
    
    slug = f"jordan-{base_slug}"
    if slug in seen_slugs:
        variant_slug = sku.lower().replace('/', '-').replace(' ', '-')
        slug = f"jordan-{base_slug}-{variant_slug}"
        slug = re.sub(r'[^a-z0-9\-]', '', slug)
        slug = re.sub(r'\-+', '-', slug).strip('-')
        
    seen_slugs.add(slug)
    
    extracted_products.append({
        "b": "Jordan Fitness",
        "n": name,
        "c": cat_name,
        "sec": sec_name,
        "sr": sr_name,
        "s": slug,
        "img": f"assets/images/products/jordan/{img_slug}.jpg",
        "d": desc[:400] + "..." if len(desc) > 400 else desc
    })

print("Extracted products count:", len(extracted_products))

by_cat = {}
for p in extracted_products:
    c = p['c']
    if c not in by_cat:
        by_cat[c] = []
    by_cat[c].append(p)

sampled_products = []
for c, prods in by_cat.items():
    print(f"Category '{c}' has {len(prods)} products.")
    max_items = 8
    if "Dumbbells" in c or "Kettlebells" in c or "Plates" in c or "Bumper" in c:
        max_items = 12
    if "Racks" in c or "Mats" in c or "BlazePod" in c:
        max_items = 7
        
    if len(prods) <= max_items:
        sampled_products.extend(prods)
    else:
        step = len(prods) / max_items
        for idx in range(max_items):
            sampled_products.append(prods[int(idx * step)])

print("Sampled products count:", len(sampled_products))

with open(APP_JS, 'r') as f:
    code = f.read()

m = re.search(r'const PRODUCTS\s*=\s*(\[.*?\]);', code, re.DOTALL)
if not m:
    raise Exception("Could not find PRODUCTS array in app.js")

all_products = json.loads(m.group(1))

# Filter out old Jordan Fitness products
all_products = [p for p in all_products if p['b'] != 'Jordan Fitness']

# Add new sampled products
all_products.extend(sampled_products)

# Serialize back
new_products_js = "const PRODUCTS = " + json.dumps(all_products) + ";"
code = code[:m.start()] + new_products_js + code[m.end():]

with open(APP_JS, 'w') as f:
    f.write(code)

print("Injected successfully!")
