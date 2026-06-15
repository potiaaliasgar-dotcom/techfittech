import openpyxl
import re

EXCEL_PATH = "/Users/batman/Desktop/techfittech/BH Price List 2026 (BH, Jordan, Bendis pilates).xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
sheet = wb['Jordan']

print("Max rows:", sheet.max_row)

products = []
for i in range(4, sheet.max_row + 1):
    sku = sheet.cell(row=i, column=1).value
    product = sheet.cell(row=i, column=3).value
    price = sheet.cell(row=i, column=4).value
    
    if not sku or not product or price is None or str(price).startswith('#'):
        continue
        
    product_clean = product.strip()
    lines = product_clean.split('\n')
    name = lines[0].strip()
    name = ' '.join(name.split())
    
    desc = ""
    if len(lines) > 1:
        desc = " ".join([l.strip() for l in lines[1:] if l.strip()])
    else:
        desc = f"Premium {name} by Jordan Fitness, designed for commercial health clubs, boutique gyms, and professional training facilities."
        
    products.append({
        'sku': sku,
        'name': name,
        'desc': desc,
        'price': price
    })

print(f"Total rows parsed: {len(products)}")

# Let's write a function to extract weight/size from product name
# e.g., "JORDAN 10kg Rubber Barbell", "JORDAN 1kg Hex Rubber Dumbbells (Pair)"
def extract_weight_or_size(name):
    # Match patterns like:
    # 10kg, 12.5kg, 1.25kg, 1kg-10kg, 6ft, 7ft, 2.5m, 0.5-5kg, 32mm, 42 pairs
    match = re.search(r'\b\d+(?:\.\d+)?\s*(?:kg|kg\b|mm|ft|m)\b', name, flags=re.IGNORECASE)
    if match:
        return match.group(0)
    # Range match: e.g. 1kg-10kg or 2.5kg-50kg
    range_match = re.search(r'\b\d+(?:\.\d+)?\s*-\s*\d+(?:\.\d+)?\s*(?:kg|kg\b|mm|ft|m)\b', name, flags=re.IGNORECASE)
    if range_match:
        return range_match.group(0)
    return None

# Let's test base names and groups
groups = {}
for p in products:
    name = p['name']
    sku = p['sku']
    
    # We clean the name to find the base name
    # Remove sizes/weights
    cleaned = re.sub(r'\b\d+(?:\.\d+)?\s*(?:kg|kg\b|mm|ft|m)\b', '', name, flags=re.IGNORECASE)
    # Remove ranges: e.g. 1kg-10kg, 1-10kg, etc.
    cleaned = re.sub(r'\b\d+(?:\.\d+)?\s*(?:kg|kg\b|mm|ft|m)?\s*-\s*\d+(?:\.\d+)?\s*(?:kg|kg\b|mm|ft|m)\b', '', cleaned, flags=re.IGNORECASE)
    # Remove specific details like "Yellow", "Blue", "Black", "Red", "Grey" if they are just colour codes
    cleaned = re.sub(r'\b(?:yellow|blue|black|red|grey|purple|green|dark red|orange|pink)\b', '', cleaned, flags=re.IGNORECASE)
    cleaned = ' '.join(cleaned.split())
    
    # Clean up punctuation at ends or extra dashes
    cleaned = re.sub(r'\s*-\s*$', '', cleaned)
    cleaned = re.sub(r'^\s*-\s*', '', cleaned)
    cleaned = cleaned.strip()
    
    if cleaned not in groups:
        groups[cleaned] = []
    groups[cleaned].append(p)

print(f"Total base groups found: {len(groups)}")
# Print groups with more than 1 item
for k, v in sorted(groups.items(), key=lambda x: len(x[1]), reverse=True)[:25]:
    print(f"\n{k} ({len(v)} variants):")
    for item in v[:5]:
        print(f"  - {item['name']} (SKU: {item['sku']})")
    if len(v) > 5:
        print(f"  - ... and {len(v)-5} more")
